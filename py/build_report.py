"""
Скрипт для построения Excel-отчёта из данных заказов.
Этапы:
1. Создание SQLite базы и загрузка в нее CSV-файлов.
2. Выполнение основного SQL-запроса (sql/export.sql) для получения данных.
3. Формирование Excel-файла с листами: Orders, Summary, Dashboard, Checks.
"""
import argparse, sqlite3, csv, sys
from pathlib import Path
import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill


def setup_database(db_path: str, data_dir: Path) -> sqlite3.Connection:
    """Создаёт и наполняет базу данных SQLite из CSV-файлов."""
    print(f"Connecting to SQLite database: {db_path}")
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode=WAL;")
    
    print("Creating database schema...")
    cursor = conn.cursor()
    cursor.executescript("""
        DROP TABLE IF EXISTS sellers;
        CREATE TABLE sellers (id INTEGER PRIMARY KEY, name TEXT NOT NULL);

        DROP TABLE IF EXISTS orders;
        CREATE TABLE orders (id INTEGER PRIMARY KEY, external_id TEXT NOT NULL, date TEXT NOT NULL, channel TEXT NOT NULL, seller_id INTEGER, status TEXT NOT NULL, updated_at TEXT NOT NULL, delivered_at TEXT, FOREIGN KEY(seller_id) REFERENCES sellers(id));

        DROP TABLE IF EXISTS order_items;
        CREATE TABLE order_items (order_id INTEGER NOT NULL, sku TEXT NOT NULL, qty INTEGER NOT NULL, revenue REAL NOT NULL, cost REAL NOT NULL, FOREIGN KEY(order_id) REFERENCES orders(id));
    """)

    # Загрузка данных
    for table_name in ['sellers', 'orders', 'order_items']:
        csv_path = data_dir / f'{table_name}.csv'
        print(f"Loading data from {csv_path.name} into '{table_name}'...")
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            header = next(reader)
            data = list(reader)
        
        placeholders = ', '.join(['?'] * len(header))
        sql = f"INSERT INTO {table_name} ({', '.join(header)}) VALUES ({placeholders})"
        cursor.executemany(sql, data)

    conn.commit()
    return conn


def execute_export_query(conn: sqlite3.Connection, sql_file: Path, days: int) -> pd.DataFrame:
    """Выполняет SQL-скрипт и возвращает результат в виде DataFrame."""
    print(f"Executing query from {sql_file.name} for the last {days} days...")
    
    with open(sql_file, 'r', encoding='utf-8') as f:
        sql_script = f.read()
    
    select_start_index = sql_script.find("WITH")
    
    # Всё, что до "WITH" - это команды создания индексов
    index_script = sql_script[:select_start_index]
    
    # Всё, что после "WITH" (включая его) - это SELECT-запрос
    select_query = sql_script[select_start_index:]
    
    if index_script.strip():
        print("Creating indexes...")
        conn.executescript(index_script)
        print("Indexes created successfully.")
    
    # А теперь выполняем SELECT через pandas
    df = pd.read_sql_query(select_query, conn, params={'days': days})
    
    print(f"Query returned {len(df)} rows.")
    return df


def build_excel_report(df: pd.DataFrame, output_file: Path):
    """Собирает и записывает итоговый Excel-отчет."""
    print(f"Building Excel report at {output_file}...")
    
    with pd.ExcelWriter(str(output_file), engine='openpyxl') as writer:
        # --- Лист 1: Orders ---
        df.to_excel(writer, sheet_name='Orders', index=False)
        print(" 'Orders' sheet created.")

        # --- Лист 2: Summary ---
        summary_df = df.groupby(['channel', 'seller']).agg(
            revenue=('revenue', 'sum'),
            cost=('cost', 'sum'),
            margin=('margin', 'sum'),
            items_count=('sku', 'count'),
            unique_orders=('order_id', 'nunique')
        ).sort_values(by='margin', ascending=False).round(2)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        print(" 'Summary' sheet created.")

        # --- Лист 3: Dashboard ---
        workbook = writer.book
        sheet = workbook.create_sheet('Dashboard')
        writer.sheets['Dashboard'] = sheet

        # Данные для графика 1: Маржа по каналам
        margin_by_channel = df.groupby('channel')['margin'].sum().round(2).reset_index()
        margin_by_channel.to_excel(writer, sheet_name='Dashboard', startrow=0, index=False)

        # Данные для графика 2: Воронка конверсии
        orders_df = df[['order_id', 'status']].drop_duplicates()
        status_order = ["created", "paid", "prod_started", "shipped", "delivered"]
        status_counts = orders_df['status'].value_counts().reindex(status_order, fill_value=0)
        
        funnel_data = []
        previous_step_count = status_counts.get("created", 1) # 1 чтобы избежать деления на 0
        
        for status in status_order:
            count = status_counts.get(status, 0)
            step_conversion = (count / previous_step_count * 100) if previous_step_count > 0 else 0
            total_conversion = (count / status_counts.get("created", 1) * 100)
            
            funnel_data.append([status, count, f"{step_conversion:.1f}%", f"{total_conversion:.1f}%"])
            previous_step_count = count # База для следующего шага - текущий шаг
        
        funnel_df = pd.DataFrame(funnel_data, columns=["Status", "Order Count", "Step Conv %", "Total Conv %"])
        funnel_df.to_excel(writer, sheet_name='Dashboard', startrow=0, startcol=4, index=False)

        # График 1: Маржа
        chart1 = BarChart()
        chart1.title = "Маржа по каналам"
        data = Reference(sheet, min_col=2, min_row=1, max_row=len(margin_by_channel) + 1)
        cats = Reference(sheet, min_col=1, min_row=2, max_row=len(margin_by_channel) + 1)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        sheet.add_chart(chart1, "A10")
        
        # График 2: Воронка
        chart2 = BarChart()
        chart2.title = "Воронка заказов по статусам"
        data = Reference(sheet, min_col=6, min_row=1, max_row=len(funnel_df) + 1)
        cats = Reference(sheet, min_col=5, min_row=2, max_row=len(funnel_df) + 1)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
        sheet.add_chart(chart2, "J10")
        print(" 'Dashboard' sheet with charts created.")

        # --- Лист 4: Checks ---
        checks_sheet = workbook.create_sheet('Checks')
        # Проверка на дубликаты
        deduped_orders_check = df[['order_id', 'external_id']].drop_duplicates()
        duplicates_in_result = deduped_orders_check[deduped_orders_check.duplicated(subset=['external_id'], keep=False)]
        
        checks = {
            "Количество <= 0": df[df['qty'] <= 0],
            "Отрицательная маржа": df[df['margin'] < 0],
            "Отсутствует продавец": df[df['seller'] == 'UNKNOWN_SELLER'],
            "Дубликаты external_id в выгрузке (должно быть 0)": df[df['order_id'].isin(duplicates_in_result['order_id'])]
        }
        
        row_offset = 1
        header_font = Font(bold=True)
        title_font = Font(bold=True, size=14)
        ok_fill = PatternFill(start_color="D6EFD6", end_color="D6EFD6", fill_type="solid")

        for title, check_df in checks.items():
            checks_sheet.cell(row=row_offset, column=1, value=f"{title} (найдено: {len(check_df)})").font = title_font
            row_offset += 2
            
            if not check_df.empty:
                headers = list(check_df.columns)
                for c_idx, col_name in enumerate(headers, 1):
                    cell = checks_sheet.cell(row=row_offset, column=c_idx, value=col_name)
                    cell.font = header_font
                    
                for _, row_data in check_df.iterrows():
                    row_offset += 1
                    for c_idx, value in enumerate(row_data, 1):
                        checks_sheet.cell(row=row_offset, column=c_idx, value=value)
                row_offset += 2
            else:
                cell = checks_sheet.cell(row=row_offset, column=1, value="Проблем не найдено")
                cell.fill = ok_fill
                row_offset += 2
        print(" 'Checks' sheet created.")


def main():
    parser = argparse.ArgumentParser(description="Сборка Excel-отчета по данным о заказах.")
    parser.add_argument("--days", type=int, default=90, help="Количество дней для отчета.")
    parser.add_argument("--out", type=Path, default="excel/Report.xlsx", help="Путь для сохранения Excel-отчета.")
    parser.add_argument("--db", default=":memory:", help="Путь к файлу SQLite или ':memory:'.")
    args = parser.parse_args()

    args.out.parent.mkdir(parents=True, exist_ok=True)
    
    conn = None
    try:
        data_dir = Path(__file__).parent.parent / 'data'
        sql_file = Path(__file__).parent.parent / 'sql' / 'export.sql'
        
        conn = setup_database(str(args.db), data_dir)
        report_df = execute_export_query(conn, sql_file, args.days)

        if report_df.empty:
            print("Предупреждение: SQL-запрос не вернул данных. Отчет будет содержать пустые листы.")
        
        build_excel_report(report_df, args.out)
        
        print("\nСкрипт успешно завершен.")
        print(f"Итоговый отчет сохранен в: {args.out}")
        return 0
        
    except Exception as e:
        print(f"\nОшибка во время выполнения скрипта: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc(file=sys.stderr)
        return 1
    finally:
        if conn:
            conn.close()
            print("Соединение с базой данных закрыто.")


if __name__ == "__main__":
    sys.exit(main())